from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import requests
from bs4 import BeautifulSoup
import openpyxl

# To fetch the ISIN number from the excel sheet.
excel_path = "H:\Python Codes\Comb_Data.xlsx"
workbook = openpyxl.load_workbook(excel_path)
sheet = workbook.active # workbook.get_sheet_by_name("Sheet1")
rows = sheet.max_row
cols = sheet.max_column

for r in range(15,21):
# Taking the ISIN value and looping it.
    isin = sheet.cell(row=r,column=2).value
# Automating to access google chrome.
    driver = webdriver.Chrome(executable_path="H:\Softwares\chromedriver_win32\chromedriver.exe")
    url = "https://markets.businessinsider.com/stocks"
# Using the url to access that particular website.
    driver.get(url)
# Searching for results using ISIN number.
    ele = driver.find_element_by_name("_search")
    ele.send_keys(isin)  # ele.send_keys("PHY0001Z1040")
    ele.send_keys(Keys.ENTER)
    search_results=driver.current_url[:49]  # print(search_results) # print(driver.current_url)
    no_search="https://markets.businessinsider.com/searchresults"
    if search_results != no_search:
        stock_url = driver.current_url
# Web-scraping to get the market capital.
        res1 = requests.get(stock_url).text
        soup1 = BeautifulSoup(res1, 'lxml')
        # mar = driver.find_element_by_class_name("price-row-price")
        try:
            market_cap = soup1.select('.price-row-price')[2].text
        except:
            market_cap = 'No Value'  # print('This is The Market Capitalization: ',market_cap)  # Market Cap is String
        driver.quit()
# Going to other url to get revenue.
        fin_url = stock_url[:-6]+'/financials' # print(fin_url)
        res2 = requests.get(fin_url).text
        soup2 = BeautifulSoup(res2, 'lxml')
        data_list = []
# Scraping among different HTML Tag classes to get revenue.
        for cl in soup2.find_all('td'):
            data_list.append(cl.text)   # print(data_list)
        try:
            index = data_list.index('Revenue')
            index += 1
            rev = data_list[index]
            rev = int(rev.replace(',', ''))
        except:
            rev = 'No Value' # print('This is the Revenue: ',rev)
        sheet.cell(row=r,column=7).value = market_cap
        sheet.cell(row=r,column=8).value = rev
        workbook.save(excel_path) # Saving those results back into excel.
    else:
# This part is executed if there are no values in the website.
        driver.quit()
        market_cap='Not Found'
        rev='Not Found'
        sheet.cell(row=r, column=7).value = market_cap
        sheet.cell(row=r, column=8).value = rev
        workbook.save(excel_path)